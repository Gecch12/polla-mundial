import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from dateutil import parser as dateparser

ROOT = Path(__file__).resolve().parents[1]
DATA_JS = ROOT / 'assets' / 'js' / 'data.js'
XLSX = ROOT / 'inbox' / 'latest.xlsx'
EMAIL_JSON = ROOT / 'inbox' / 'email.json'
PROCESSED = ROOT / 'inbox' / 'processed_message_ids.txt'


def log(message: str) -> None:
    print(f"[polla-debug] {datetime.utcnow().isoformat(timespec='seconds')}Z | {message}", flush=True)


def load_current() -> Dict[str, Any]:
    txt = DATA_JS.read_text(encoding='utf-8')
    m = re.match(r'\s*window\.POLLA_DATA\s*=\s*(.*);\s*$', txt, re.S)
    if not m:
        raise ValueError('assets/js/data.js does not match expected window.POLLA_DATA format')
    return json.loads(m.group(1))


def save_data(data: Dict[str, Any]) -> None:
    DATA_JS.write_text('window.POLLA_DATA = ' + json.dumps(data, ensure_ascii=False, separators=(',', ':')) + ';\n', encoding='utf-8')


def normalize_name(x: Any) -> str:
    s = str(x or '').strip()
    s = re.sub(r'\s+', ' ', s)
    return s.upper()


def to_number(x: Any) -> Optional[float]:
    if pd.isna(x):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(',', '.')
    m = re.search(r'-?\d+(?:\.\d+)?', s)
    return float(m.group()) if m else None


def infer_date(email: Dict[str, Any]) -> str:
    subject = email.get('subject', '')
    sent = email.get('sent', '')
    # Prefer date mentioned in subject/body if present: 29 JUNIO 2026, 29/06/2026, etc.
    month_map = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
        'julio': 7, 'agosto': 8, 'setiembre': 9, 'septiembre': 9, 'octubre': 10,
        'noviembre': 11, 'diciembre': 12,
        'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6, 'jul': 7,
        'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12,
    }
    text = (subject + '\n' + email.get('bodyText', ''))[:5000].lower()
    m = re.search(r'\b(\d{1,2})\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|setiembre|septiembre|octubre|noviembre|diciembre|ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)\s+(20\d{2})\b', text)
    if m:
        return f"{int(m.group(3)):04d}-{month_map[m.group(2)]:02d}-{int(m.group(1)):02d}"
    m = re.search(r'\b(\d{1,2})[/-](\d{1,2})[/-](20\d{2})\b', text)
    if m:
        return f"{int(m.group(3)):04d}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    if sent:
        try:
            d = dateparser.parse(sent, dayfirst=True, fuzzy=True)
            # Beto often sends next morning; subject date is better, but if not found use send date.
            return d.date().isoformat()
        except Exception:
            pass
    return datetime.utcnow().date().isoformat()


def find_ranking_table(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Extract the official ranking table from Beto's PUNTAJES sheet.

    Production/debug version:
    - opens workbook in read_only mode (important for Beto's Excel);
    - scans bounded ranges only;
    - detects the right-side ranking table with several possible layouts:
        A: TOTAL | POS | NAME
        B: TOTAL | blank | POS | NAME
        C: TOTAL | blank | blank | POS | NAME  (layout seen in latest.xlsx)
    - refuses to publish if validation looks wrong.
    """
    from openpyxl import load_workbook

    log(f"Abriendo Excel en read_only: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    log(f"Excel abierto. Hojas: {', '.join(wb.sheetnames)}")

    preferred = [name for name in wb.sheetnames if normalize_name(name) == 'PUNTAJES']
    preferred += [name for name in wb.sheetnames if 'PUNTAJES' in normalize_name(name) and name not in preferred]
    sheets = preferred + [n for n in wb.sheetnames if n not in preferred]
    log(f"Orden de búsqueda: {', '.join(sheets)}")

    all_candidates: List[Dict[str, Any]] = []

    def valid_name(v: Any) -> str:
        name = normalize_name(v)
        if len(name) < 3:
            return ''
        bad = [
            'TOTAL', 'PUNTOS', 'PUESTOS', 'PARTICIPANTE', 'CLASIF', 'ETAPA',
            'GRUPO', 'REGLAMENTO', 'FECHA', 'FINAL', 'OCTAVOS', 'DIECISEISAVOS',
            'PERU', 'MUNDIAL', 'PUESTO', 'LOCAL', 'VISITA', 'GANADOR', 'PARR'
        ]
        if any(b in name for b in bad):
            return ''
        if not re.search(r'[A-ZÁÉÍÓÚÑ]', name):
            return ''
        if re.search(r'https?://|@', name, re.I):
            return ''
        # Avoid names that are mostly numbers/single codes.
        if len(re.sub(r'[^A-ZÁÉÍÓÚÑ ]', '', name)) < 3:
            return ''
        return name

    def get_grid(ws, max_rows: int, max_cols: int) -> List[List[Any]]:
        log(f"Leyendo hoja '{ws.title}' en memoria: rows={max_rows}, cols={max_cols}")
        grid: List[List[Any]] = []
        for row in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=True):
            grid.append(list(row))
        log(f"Hoja '{ws.title}' cargada: {len(grid)} filas")
        return grid

    def cell(grid: List[List[Any]], r: int, c: int) -> Any:
        if r < 0 or c < 0 or r >= len(grid) or c >= len(grid[r]):
            return None
        return grid[r][c]

    def scan_pattern(grid: List[List[Any]], max_row: int, max_col: int, total_offset: int, pos_offset: int, name_offset: int, label: str, sh_name: str) -> None:
        # total_offset is always 0, but kept explicit for readability.
        for total_col in range(max(0, max_col - 90), max_col - max(pos_offset, name_offset) - 1):
            rows: List[Dict[str, Any]] = []
            for r in range(max_row):
                pts = to_number(cell(grid, r, total_col + total_offset))
                pos = to_number(cell(grid, r, total_col + pos_offset))
                name = valid_name(cell(grid, r, total_col + name_offset))
                if pts is None or pos is None or not name:
                    continue
                if not (1 <= int(pos) <= 250):
                    continue
                if not (120 <= int(pts) <= 1000):
                    continue
                rows.append({'Participante': name, 'Puntos': int(round(pts)), 'Posicion': int(round(pos))})
            if len(rows) >= 50:
                rows.sort(key=lambda r: (int(r['Posicion']), -int(r['Puntos'])))
                leader = rows[0]
                # Prefer sheets literally named PUNTAJES and layouts with the right-side leader.
                sheet_bonus = 10000 if normalize_name(sh_name) == 'PUNTAJES' else 5000 if 'PUNTAJES' in normalize_name(sh_name) else 0
                leader_bonus = 5000 if leader['Posicion'] == 1 and leader['Puntos'] >= 150 else 0
                score = sheet_bonus + leader_bonus + len(rows) * 1000 + int(leader['Puntos'])
                log(f"Candidato {label} en '{sh_name}' col {total_col+1}: {len(rows)} filas, top={leader['Posicion']}. {leader['Participante']} {leader['Puntos']}, score={score}")
                all_candidates.append({'score': score, 'rows': rows, 'sheet': sh_name, 'pattern': f'{label} total_col={total_col+1}'})

    for sh_name in sheets:
        ws = wb[sh_name]
        raw_max_row = ws.max_row or 0
        raw_max_col = ws.max_column or 0
        max_row = min(raw_max_row, 350)
        max_col = min(raw_max_col, 220)
        log(f"Escaneando hoja '{sh_name}': max_row={raw_max_row}, max_col={raw_max_col}, bounded={max_row}x{max_col}")
        if max_row < 10 or max_col < 5:
            log(f"Saltando hoja '{sh_name}' por tamaño insuficiente")
            continue

        grid = get_grid(ws, max_row, max_col)

        log(f"Buscando patrón A TOTAL|POS|NOMBRE en '{sh_name}'")
        scan_pattern(grid, max_row, max_col, 0, 1, 2, 'A TOTAL|POS|NOMBRE', sh_name)

        log(f"Buscando patrón B TOTAL|blank|POS|NOMBRE en '{sh_name}'")
        scan_pattern(grid, max_row, max_col, 0, 2, 3, 'B TOTAL|blank|POS|NOMBRE', sh_name)

        log(f"Buscando patrón C TOTAL|blank|blank|POS|NOMBRE en '{sh_name}'")
        scan_pattern(grid, max_row, max_col, 0, 3, 4, 'C TOTAL|blank|blank|POS|NOMBRE', sh_name)

    log(f"Candidatos encontrados: {len(all_candidates)}")
    if not all_candidates:
        raise ValueError('No pude encontrar la tabla oficial de ranking en la hoja PUNTAJES. No se publica nada.')

    chosen = max(all_candidates, key=lambda x: x['score'])
    rows = chosen['rows']
    log(f"Candidato elegido: hoja={chosen['sheet']}, patrón={chosen['pattern']}, filas_raw={len(rows)}")

    dedup: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        name = row['Participante']
        if name not in dedup or row['Puntos'] > dedup[name]['Puntos']:
            dedup[name] = row
    rows = list(dedup.values())
    rows.sort(key=lambda r: (int(r.get('Posicion') or 9999), -int(r['Puntos']), r['Participante']))

    log(f"Después de deduplicar: {len(rows)} participantes")
    if rows:
        log("Top extraído: " + " | ".join(f"{r['Posicion']}. {r['Participante']} {r['Puntos']}" for r in rows[:5]))

    if len(rows) < 70:
        raise ValueError(f'Validación fallida: solo se extrajeron {len(rows)} participantes. No se publica.')
    if rows[0]['Puntos'] < 150:
        raise ValueError(f'Validación fallida: líder con {rows[0]["Puntos"]} puntos parece incorrecto. No se publica.')
    if rows[0]['Participante'] in {'TERCEROS', 'TOTAL', 'PUNTOS'}:
        raise ValueError('Validación fallida: el líder extraído no es un participante real. No se publica.')
    if rows[0]['Posicion'] != 1:
        raise ValueError(f'Validación fallida: primer registro no tiene posición 1 ({rows[0]}). No se publica.')

    log(f"Ranking oficial OK: {len(rows)} participantes. Líder {rows[0]['Participante']} {rows[0]['Puntos']}.")
    return rows

def compute_new_snapshot(current: Dict[str, Any], latest_rows: List[Dict[str, Any]], date_str: str) -> None:
    base = current.get('base', [])
    # Remove same date to make reruns idempotent.
    base = [r for r in base if r.get('Fecha') != date_str]
    previous_dates = sorted({r.get('Fecha') for r in base if r.get('Fecha') and r.get('Fecha') < date_str})
    prev_date = previous_dates[-1] if previous_dates else None
    prev_by_name = {normalize_name(r.get('Participante')): r for r in base if r.get('Fecha') == prev_date} if prev_date else {}

    leader_points = max(r['Puntos'] for r in latest_rows)
    new_rows = []
    for row in latest_rows:
        prev = prev_by_name.get(row['Participante'])
        puntos_dia = None if prev is None else row['Puntos'] - int(prev.get('Puntos') or 0)
        cambio = None if prev is None else int(prev.get('Posicion') or row['Posicion']) - row['Posicion']
        new_rows.append({
            'Fecha': date_str,
            'Participante': row['Participante'],
            'Posicion': row['Posicion'],
            'Puntos': row['Puntos'],
            'Puntos_Dia': puntos_dia,
            'Distancia_Lider': leader_points - row['Puntos'],
            'Cambio_Posicion': cambio,
        })

    current['base'] = sorted(base + new_rows, key=lambda r: (r.get('Fecha',''), int(r.get('Posicion') or 9999)))
    rebuild_derived(current)


def rebuild_derived(data: Dict[str, Any]) -> None:
    base = data.get('base', [])
    dates = sorted({r['Fecha'] for r in base})
    movers = []
    leader_days: Dict[str, int] = {}
    for d in dates:
        rows = [r for r in base if r['Fecha'] == d]
        if not rows:
            continue
        leader = min(rows, key=lambda r: int(r.get('Posicion') or 9999))
        leader_days[leader['Participante']] = leader_days.get(leader['Participante'], 0) + 1
        changes = [r for r in rows if r.get('Cambio_Posicion') is not None]
        up = max(changes, key=lambda r: r['Cambio_Posicion'], default=None)
        down = min(changes, key=lambda r: r['Cambio_Posicion'], default=None)
        if up or down:
            movers.append({
                'Fecha': d,
                'Mayor_Subida': up['Participante'] if up else None,
                'Puestos_Subidos': up['Cambio_Posicion'] if up else None,
                'Mayor_Bajada': down['Participante'] if down else None,
                'Puestos_Bajados': down['Cambio_Posicion'] if down else None,
                'Lider': leader['Participante'],
                'Puntos_Lider': leader['Puntos'],
            })
    data['movers'] = movers
    data['leaders'] = sorted([
        {'Participante': k, 'Dias_Lider': v} for k,v in leader_days.items()
    ], key=lambda r: (-r['Dias_Lider'], r['Participante']))

    latest_date = dates[-1] if dates else None
    latest = [r for r in base if r['Fecha'] == latest_date]
    first_date = dates[0] if dates else None
    first = {r['Participante']: r for r in base if r['Fecha'] == first_date}
    if latest:
        leader = min(latest, key=lambda r: int(r.get('Posicion') or 9999))
        climbs = []
        for r in latest:
            f = first.get(r['Participante'])
            if f:
                climbs.append({'name': r['Participante'], 'positions': int(f['Posicion']) - int(r['Posicion'])})
        biggest_up = max(climbs, key=lambda x: x['positions'], default={'name': None, 'positions': 0})
        biggest_down = min(climbs, key=lambda x: x['positions'], default={'name': None, 'positions': 0})
        data['summary'] = {
            'updated': latest_date,
            'leader': leader['Participante'],
            'leader_points': leader['Puntos'],
            'participants': len(latest),
            'days': len(dates),
            'biggest_total_climb': biggest_up,
            'biggest_total_drop': biggest_down,
        }


def update_email(data: Dict[str, Any], email: Dict[str, Any]) -> None:
    data['email'] = {
        'subject': email.get('subject', ''),
        'from': email.get('from', ''),
        'sent': email.get('sent', ''),
        'attachment': email.get('attachment', ''),
        'body': email.get('bodyText', ''),
    }


def main() -> None:
    log('Inicio update_from_email.py')
    if not XLSX.exists():
        raise FileNotFoundError('Missing inbox/latest.xlsx')
    email = json.loads(EMAIL_JSON.read_text(encoding='utf-8')) if EMAIL_JSON.exists() else {}
    msg_id = email.get('messageId', '')
    processed = set(PROCESSED.read_text(encoding='utf-8').splitlines()) if PROCESSED.exists() else set()
    # Idempotency: if data.js already has this email subject/attachment, allow no-op safely.

    data = load_current()
    latest_date = infer_date(email)
    latest_rows = find_ranking_table(XLSX)
    log('Ranking extraído correctamente')
    compute_new_snapshot(data, latest_rows, latest_date)
    update_email(data, email)
    save_data(data)

    if msg_id and msg_id not in processed:
        with PROCESSED.open('a', encoding='utf-8') as f:
            f.write(msg_id + '\n')

    print(f'Updated Polla data for {latest_date}: {len(latest_rows)} participants, leader {data["summary"]["leader"]} ({data["summary"]["leader_points"]} pts)')


if __name__ == '__main__':
    main()
